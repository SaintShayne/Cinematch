#!/usr/bin/env python3
"""
CineMatch QA Test Plan — Excel generator (v2).
4 sheets: Dashboard | Manual Test Cases | Automation Test Cases | Automation Setup Guide
Automation tool: Playwright (definitive).
"""

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ── Style helpers ──────────────────────────────────────────────────────────────

def F(h): return PatternFill("solid", fgColor=h)
def Fn(bold=False, color="1A1A2E", size=9, name="Calibri"):
    return Font(bold=bold, color=color, size=size, name=name)
def Al(h="left", v="top", wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
def Bd(color="C5D9EE"):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)
def Bd_med(color="9AB8D8"):
    m = Side(style="medium", color=color)
    t = Side(style="thin",   color=color)
    return Border(left=m, right=m, top=m, bottom=t)

# ── Palette ────────────────────────────────────────────────────────────────────
HDR_BG   = "1F4E79"; HDR_FG   = "FFFFFF"
PH_BG    = "D6E4F0"; PH_FG    = "1F4E79"
ODD_BG   = "FFFFFF"; EVEN_BG  = "EEF4FB"
EMPTY_BG = "F5F8FA"
AUTO_HDR = "145A32"; AUTO_FG  = "FFFFFF"   # green header for automation sheet
AUTO_PH  = "D5F5E3"; AUTO_PFG = "145A32"
CRIT_BG  = "FFD0D0"; CRIT_FG  = "922B21"
HIGH_BG  = "FFEAD6"; HIGH_FG  = "943126"
MED_BG   = "FFFAD6"; MED_FG   = "7D6608"
LOW_BG   = "D5F5E3"; LOW_FG   = "1E8449"
PW_BG    = "E8F4FD"; PW_FG    = "1A5276"   # Playwright tool badge

PRIO_BG = {"Critical": CRIT_BG, "High": HIGH_BG, "Medium": MED_BG, "Low": LOW_BG}
PRIO_FG = {"Critical": CRIT_FG, "High": HIGH_FG, "Medium": MED_FG, "Low": LOW_FG}

# ── Phase labels ───────────────────────────────────────────────────────────────
PHASES = {
    "0":   "Phase 0   — Environment Setup",
    "1":   "Phase 1   — Backend Health Check",
    "2":   "Phase 2   — Homepage",
    "3A":  "Phase 3A  — Auth: Logged-Out State",
    "3B":  "Phase 3B  — Auth: Email Registration",
    "3C":  "Phase 3C  — Auth: Email Login",
    "3D":  "Phase 3D  — Auth: Google OAuth",
    "4":   "Phase 4   — Browse",
    "5":   "Phase 5   — Movie Detail Page",
    "6":   "Phase 6   — People Page",
    "7":   "Phase 7   — Search",
    "8":   "Phase 8   — Recommendations",
    "9":   "Phase 9   — Watchlist",
    "10":  "Phase 10  — History",
    "11":  "Phase 11  — Chat (AI Assistant)",
    "12A": "Phase 12A — Support: Logged-Out Guard",
    "12B": "Phase 12B — Support: Local Currency Display",
    "12C": "Phase 12C — Support: Tier Feature List",
    "12D": "Phase 12D — Support: Stripe Checkout",
    "12E": "Phase 12E — Support: Supporter Tag",
    "13":  "Phase 13  — Report an Issue",
    "13A": "Phase 13A — Report: File Attachment",
    "14A": "Phase 14A — Admin: Login",
    "14B": "Phase 14B — Admin: Dashboard",
    "14C": "Phase 14C — Admin: Feature Flags",
    "14D": "Phase 14D — Admin: 2FA (Telegram)",
    "15":  "Phase 15  — About Page",
    "16":  "Phase 16  — Edge Cases",
    "17":  "Phase 17  — Cross-Browser Smoke Test",
}

# ══════════════════════════════════════════════════════════════════════════════
# MANUAL TEST CASES
# [tc_id, phase, category, title, priority, preconditions, steps, expected]
# ══════════════════════════════════════════════════════════════════════════════
MANUAL = [
    # ── Phase 0 — Environment Setup ───────────────────────────────────────────
    ["TC-001","0","Setup","Open Chrome in normal mode (not Incognito)","Low",
     "Chrome installed",
     "1. Open Google Chrome.\n2. Confirm window is NOT Incognito.\n3. Confirm cookies allowed for your-app.vercel.app.",
     "Chrome open in normal mode; cookies enabled; session data will persist"],

    ["TC-002","0","Setup","Enable DevTools Network tab with Preserve Log","Low",
     "Chrome open",
     "1. Press F12.\n2. Click Network tab.\n3. Tick 'Preserve log' checkbox.\n4. Keep DevTools open throughout testing.",
     "Network tab active; Preserve log ticked; all HTTP events captured across navigations"],

    ["TC-003","0","Setup","Monitor DevTools Console for JavaScript errors","Low",
     "DevTools open",
     "1. Click Console tab.\n2. Click 'Clear console' (bin icon).\n3. Keep visible alongside Network tab.",
     "Console tab open and clear; any red JS errors during testing will be visible"],

    ["TC-004","0","Setup","Prepare logged-out browser for auth guard tests","Low",
     "Firefox OR Chrome Incognito available",
     "1. Open Firefox OR Chrome Incognito (Ctrl+Shift+N).\n2. Do NOT log in.\n3. Keep available for Phase 3A and 12A.",
     "Separate browser with no active CineMatch session; ready for logged-out tests"],

    ["TC-005","0","Setup","Gather Stripe test card details","Low",
     "None",
     "1. Note test card number: 4242 4242 4242 4242.\n2. Expiry: any future date (e.g. 12/30).\n3. CVC: any 3 digits (e.g. 123).\n4. ZIP: any 5 digits (e.g. 10001).",
     "Test card details ready for Phase 12D; these complete payment in Stripe test mode — no real charge"],

    ["TC-006","0","Setup","Confirm Supabase dashboard access to profiles table","Medium",
     "Supabase account with CineMatch project access",
     "1. Log into supabase.com.\n2. Open CineMatch project.\n3. Table Editor → profiles.\n4. Confirm is_supporter column exists (boolean).",
     "profiles table visible; is_supporter column present; ready to verify post-payment supporter tag"],

    ["TC-007","0","Setup","Confirm Telegram is accessible for verification","Medium",
     "Telegram account linked to CineMatch bot",
     "1. Open Telegram (desktop or mobile).\n2. Open the CineMatch bot chat.\n3. Confirm chat is active.",
     "Telegram ready; will receive forwarded reports (Phase 13) and 2FA OTPs (Phase 14D)"],

    # ── Phase 3B — Email Registration ─────────────────────────────────────────
    ["TC-023","3B","Auth","Register with a new valid email address","Critical",
     "A new email address not previously registered; logged out",
     "1. Navigate to /register.\n2. Fill all fields with valid data.\n3. Use a real email you can check.\n4. Submit.",
     "Registration accepted; 'Check your email' message or confirmation sent notification shown"],

    ["TC-024","3B","Auth","Confirmation email arrives from noreply@your-domain.com","Critical",
     "TC-023 passed; inbox accessible",
     "1. Open inbox for the email used in TC-023.\n2. Wait up to 3 minutes.\n3. Find CineMatch email.\n4. Check From address.",
     "Email arrives within 3 min; From = noreply@your-domain.com (Resend SMTP); subject references confirmation"],

    ["TC-025","3B","Auth","Confirmation link in email logs user into site","Critical",
     "TC-024 passed; email open",
     "1. Click the confirmation link in the email.\n2. Observe redirect and login state.",
     "Redirected to your-app.vercel.app; user automatically logged in; nav shows avatar"],

    # ── Phase 3D — Google OAuth ────────────────────────────────────────────────
    ["TC-029","3D","Auth","Google OAuth button triggers Google auth flow","Critical",
     "Logged out; Google account available",
     "1. Navigate to /login.\n2. Click 'Continue with Google'.\n3. Observe result.",
     "Google OAuth popup or redirect appears; Google account selection shown with CineMatch as requesting app"],

    ["TC-030","3D","Auth","Completing Google OAuth returns user fully logged in","Critical",
     "TC-029 in progress; Google account selected",
     "1. Select/confirm Google account.\n2. Grant permissions if prompted.\n3. Wait for redirect.",
     "Redirected to your-app.vercel.app; user logged in; nav shows Google account avatar"],

    ["TC-031","3D","Auth","Google OAuth user profile exists in Supabase profiles","High",
     "TC-030 passed; Supabase dashboard open",
     "1. Open Supabase → Table Editor → profiles.\n2. Search for Google account email.\n3. Inspect the row.",
     "Row exists in profiles for the Google account; created_at is recent"],

    # ── Phase 12B — Currency Display ──────────────────────────────────────────
    ["TC-063","12B","Support","Support page displays price in local currency","High",
     "Logged in; not on VPN; connection from Malaysia",
     "1. Navigate to /support.\n2. Find price on Supporter tier card.\n3. Read currency code and amount.",
     "Price shown in local currency (e.g. MYR 14 for Malaysia); NOT plain 'USD 3.00'"],

    ["TC-064","12B","Support","Currency amount uses Math.ceil — no decimal places","Medium",
     "TC-063 passed",
     "1. Note displayed local currency amount.\n2. Calculate: Math.ceil(3 × exchange rate).\n3. Compare to displayed value.",
     "Displayed amount is a whole integer; no decimal places; matches ceiling calculation"],

    ["TC-065","12B","Support","Fallback to USD 3.00 if currency detection fails","Medium",
     "DevTools open; ability to block network requests",
     "1. DevTools → Network → Request Blocking → block 'ipapi.co'.\n2. Reload /support.\n3. Observe price.",
     "Price falls back gracefully to 'USD 3.00'; no broken UI or error crash"],

    # ── Phase 12C — Tier Features ─────────────────────────────────────────────
    ["TC-066","12C","Support","Two tier cards visible: Free and Supporter","High",
     "Logged in; on /support",
     "1. Navigate to /support.\n2. Count the tier cards displayed.",
     "Exactly two tier cards: one labelled 'Free', one labelled 'Supporter'"],

    ["TC-067","12C","Support","Both tiers share the SHARED_FEATURES base list","High",
     "TC-066 passed",
     "1. Read feature list on Free tier.\n2. Read feature list on Supporter tier.\n3. Compare overlapping items.",
     "All base features appear in both tiers; nothing missing from either"],

    ["TC-068","12C","Support","Supporter tier shows unique streaming hint perk","Medium",
     "TC-066 passed",
     "1. Read Supporter tier feature list.\n2. Look for streaming-related or 'coming soon' perk.",
     "At least one unique perk appears ONLY on Supporter tier"],

    ["TC-069","12C","Support","No 'Current plan' badge on Supporter for non-supporters","High",
     "Logged in as user with is_supporter = false",
     "1. Navigate to /support as a non-supporter.\n2. Inspect Supporter tier card for badges.",
     "No 'Current plan' or 'Active' badge on Supporter tier; confirms v5 bug fix"],

    # ── Phase 12D — Stripe Checkout ───────────────────────────────────────────
    ["TC-070","12D","Support","Checkout button redirects to Stripe-hosted checkout","Critical",
     "Logged in; STRIPE_SECRET_KEY configured; Network tab open",
     "1. Click 'Support CineMatch' button.\n2. Monitor Network tab for POST /create-checkout-session.\n3. Observe URL.",
     "Redirected to checkout.stripe.com; Stripe page loads; POST returned a session URL"],

    ["TC-071","12D","Support","Stripe page shows product name 'CineMatch Supporter'","High",
     "TC-070 passed; on Stripe checkout page",
     "1. Read the product name on Stripe checkout page.",
     "Product name: 'CineMatch Supporter'; description: 'A one-time contribution...'"],

    ["TC-072","12D","Support","Stripe page shows price of $3.00 USD","High",
     "TC-070 passed; on Stripe checkout page",
     "1. Read the price on Stripe checkout page.",
     "Price shows $3.00 USD"],

    ["TC-073","12D","Support","User email is pre-filled on Stripe checkout form","High",
     "TC-070 passed; logged in with known email",
     "1. Observe the email field on Stripe checkout form.",
     "Email pre-populated with logged-in user's email; confirms customer_email passed from backend"],

    ["TC-074","12D","Support","Test card 4242... completes payment successfully","Critical",
     "TC-070-073 passed; Stripe test card from TC-005",
     "1. Enter card: 4242 4242 4242 4242.\n2. Expiry: 12/30.\n3. CVC: 123.\n4. ZIP: 10001.\n5. Click Pay.",
     "Payment accepted by Stripe test mode; no decline; redirect to success begins"],

    ["TC-075","12D","Support","Successful payment redirects back to site with success UI","Critical",
     "TC-074 passed",
     "1. Wait for Stripe redirect after payment.\n2. Observe destination URL and page content.",
     "Redirected to your-app.vercel.app; success message displayed"],

    # ── Phase 12E — Supporter Tag ─────────────────────────────────────────────
    ["TC-076","12E","Support","is_supporter = true in Supabase after payment","Critical",
     "TC-074 passed; Supabase dashboard open; webhook active on Render",
     "1. Open Supabase → Table Editor → profiles.\n2. Find your user row.\n3. Read is_supporter value.",
     "is_supporter = true; confirms POST /stripe-webhook patched the profile correctly"],

    ["TC-077","12E","Support","Support page reflects supporter status post-payment","High",
     "TC-076 confirmed; logged in as supporter",
     "1. Navigate to /support.\n2. Inspect Supporter tier card.",
     "Supporter status visually shown (badge or checkout button hidden)"],

    ["TC-078","12E","Support","Confirmed supporter not prompted to pay again","Medium",
     "TC-077 passed; logged in as confirmed supporter",
     "1. Navigate to /support.\n2. Look for checkout/Pay button.",
     "Checkout button hidden or replaced with thank-you message; no double-payment possible"],

    # ── Phase 13 — Telegram Verification ─────────────────────────────────────
    ["TC-083","13","Report","Submitted report arrives in Telegram as formatted text","Critical",
     "TC-081 passed (auto); Telegram open (TC-007)",
     "1. Check Telegram bot chat after TC-081 auto-run.\n2. Wait up to 30 seconds.",
     "Telegram message received with: category (Bug), subject containing 'TC-081', description; Markdown formatted"],

    # ── Phase 13A — Telegram Image ────────────────────────────────────────────
    ["TC-085","13A","Report","Telegram receives image attachment as inline photo","High",
     "TC-084 passed (auto); Telegram open",
     "1. Check Telegram bot chat after TC-084 auto-run.",
     "Telegram shows photo message (inline preview); caption has category, subject, description"],

    # ── Phase 14D — Telegram 2FA ──────────────────────────────────────────────
    ["TC-098","14D","Admin","Trigger Telegram OTP from admin 2FA panel","High",
     "TELEGRAM_BOT_TOKEN and TELEGRAM_CHAT_ID set in Render env vars; admin logged in",
     "1. In admin 2FA section, click 'Send OTP via Telegram'.\n2. Check Network tab for POST /admin/2fa/telegram/send.",
     "POST returns 200 with message 'OTP sent via Telegram'; expires_in_seconds: 120"],

    ["TC-099","14D","Admin","OTP code arrives in Telegram within 15 seconds","High",
     "TC-098 passed",
     "1. Check Telegram bot chat immediately.\n2. Wait up to 15 seconds.",
     "6-digit OTP arrives; message header reads 'CineMatch admin 2FA code'"],

    ["TC-100","14D","Admin","Correct OTP verifies and shows confirmation","High",
     "TC-099 passed; OTP code in hand",
     "1. Enter 6-digit OTP into admin panel verification input.\n2. Click Verify within 2 minutes.",
     "'Telegram OTP verified — 2FA enabled successfully' shown; no error"],

    # ── Phase 15 — About ──────────────────────────────────────────────────────
    ["TC-101","15","About","About page renders without errors","Medium",
     "Logged in or out",
     "1. Navigate to /about.\n2. Wait for full load.\n3. Check Console for red errors.",
     "Page renders; no red console errors; all text visible"],

    ["TC-102","15","About","No em-dash characters (--) anywhere on About page","High",
     "TC-101 passed",
     "1. Press Ctrl+F in browser.\n2. Search for em-dash character: --\n3. Also read all text visually.",
     "Zero em-dash matches; all punctuation uses natural alternatives; confirms v5 sweep"],

    ["TC-103","15","About","All links on About page are functional","Medium",
     "TC-101 passed",
     "1. Click every link on About page.\n2. Verify each goes to a valid destination.\n3. Press Back after each.",
     "All links navigate correctly; no 404 pages; no dead links"],

    # ── Phase 16 — Visual/Manual Edge Cases ───────────────────────────────────
    ["TC-105","16","Edge Cases","Slow 3G throttle shows loading states, no infinite spinner","Medium",
     "DevTools open",
     "1. DevTools → Network → throttling → select 'Slow 3G'.\n2. Reload homepage.\n3. Watch loading behaviour.",
     "Skeleton/spinner shows while loading; all spinners resolve once content arrives; no infinite loop"],

    ["TC-106","16","Edge Cases","375px mobile width shows responsive layout","High",
     "DevTools open",
     "1. Press Ctrl+Shift+M to toggle device toolbar.\n2. Set width to 375px.\n3. Navigate through Homepage, Browse, Support.\n4. Check layout on each.",
     "Layout adapts; sidebar collapses or hamburger menu appears; text readable; no overflow"],

    ["TC-107","16","Edge Cases","No horizontal scrollbar at 375px mobile width","Medium",
     "TC-106 in progress (375px active)",
     "1. At 375px, look for horizontal scrollbar at browser bottom.\n2. Try scrolling horizontally.",
     "No horizontal scrollbar; content does not overflow beyond 375px viewport"],
]

# ══════════════════════════════════════════════════════════════════════════════
# AUTOMATION TEST CASES (Playwright)
# [tc_id, phase, category, title, tool, test_file, test_data, setup,
#  what_script_does, expected, verify_accuracy, cleanup]
# ══════════════════════════════════════════════════════════════════════════════
AUTO = [
    # ── Phase 1 — Backend Health (Playwright API — no browser) ────────────────
    ["TC-008","1","Backend","GET /health returns healthy status",
     "Playwright API","tests/e2e/health.spec.js",
     "None",
     "Backend may cold-start (40s). The test has a 45s timeout set.",
     "1. Send GET to /health via APIRequestContext.\n2. Assert HTTP status === 200.\n3. Assert response.success === true.\n4. Assert response.status === 'healthy'.",
     "HTTP 200; {success:true, status:'healthy'}",
     "INTENTIONAL FAIL: Change assertion to expect status === 'broken'. Run. Confirm error: \"Expected 'broken' but received 'healthy'\". Revert.",
     "None — read-only call"],

    ["TC-009","1","Backend","GET / returns version 4.0.0",
     "Playwright API","tests/e2e/health.spec.js",
     "None",
     "Backend awake (TC-008 ran first in same file)",
     "1. Send GET to /.\n2. Assert HTTP 200.\n3. Assert response.version === '4.0.0'.\n4. Assert response.message includes 'CineMatch'.",
     "HTTP 200; version: '4.0.0'",
     "INTENTIONAL FAIL: Change '4.0.0' to '3.0.0'. Run. Confirm assertion error. Revert.",
     "None"],

    ["TC-010","1","Backend","GET /stats returns non-zero movie and genre counts",
     "Playwright API","tests/e2e/health.spec.js",
     "None",
     "Backend awake",
     "1. Send GET to /stats.\n2. Assert HTTP 200.\n3. Assert stats.total_movies > 0.\n4. Assert stats.total_genres > 0.",
     "total_movies > 0 and total_genres > 0",
     "INTENTIONAL FAIL: Change .above(0) to .above(999999). Confirm assertion fails. Revert.",
     "None"],

    # ── Phase 2 — Homepage ────────────────────────────────────────────────────
    ["TC-011","2","Homepage","Homepage loads with zero real JavaScript errors",
     "Playwright E2E","tests/e2e/homepage.spec.js",
     "None",
     "Fresh browser context; no auth required",
     "1. Register console error listener before navigation.\n2. Navigate to / and wait for networkidle.\n3. Filter out known favicon 404.\n4. Assert remaining errors array is empty.",
     "Zero real JS errors in console",
     "WATCH: Run --headed. Open DevTools inside Playwright browser. Confirm console is clean.\nINTENTIONAL FAIL: Inject page.evaluate(() => console.error('INJECTED')) before assertion. Confirm test catches it. Remove injection.",
     "None"],

    ["TC-012","2","Homepage","Trending grid shows at least 4 movie cards",
     "Playwright E2E","tests/e2e/homepage.spec.js",
     "None",
     "None",
     "1. Navigate to /.\n2. Wait for networkidle.\n3. Count movie card elements on page.\n4. Assert count >= 4.",
     "At least 4 movie cards visible",
     "INTENTIONAL FAIL: Change assertion to >= 100. Confirm failure. Revert.",
     "None"],

    ["TC-013","2","Homepage","Movie card shows poster image, title text, and year",
     "Playwright E2E","tests/e2e/homepage.spec.js",
     "None",
     "None",
     "1. Navigate to /.\n2. Locate first movie card.\n3. Assert image element is visible.\n4. Assert title text element is visible.\n5. Assert year text is visible and matches /\\d{4}/.",
     "All 3 elements visible on first card",
     "INTENTIONAL FAIL: Change year regex to /\\d{6}/. Confirm fail. Revert.",
     "None"],

    ["TC-014","2","Homepage","Clicking a movie card navigates to /movies/[slug]",
     "Playwright E2E","tests/e2e/homepage.spec.js",
     "None",
     "None",
     "1. Navigate to /.\n2. Click first movie card.\n3. Wait for networkidle.\n4. Assert page.url() contains '/movies/'.",
     "URL contains /movies/ after click",
     "INTENTIONAL FAIL: Assert URL contains '/wrong/'. Confirm fail. Revert.",
     "None"],

    ["TC-015","2","Homepage","Sidebar shows all 8 required navigation labels",
     "Playwright E2E","tests/e2e/homepage.spec.js",
     "None",
     "None",
     "1. Navigate to /.\n2. For each of [Browse, Recommendations, Watchlist, History, About, Support, Report an Issue]: assert getByRole('link',{name}) is visible.",
     "All 8 nav labels present and visible",
     "INTENTIONAL FAIL: Add 'NonExistentPage' to the check list. Confirm test fails on that item. Remove it.",
     "None"],

    ["TC-016","2","Homepage","Nav shows 'Recommendations', NOT 'For You'",
     "Playwright E2E","tests/e2e/homepage.spec.js",
     "None",
     "None",
     "1. Navigate to /.\n2. Assert link with name 'Recommendations' is visible.\n3. Assert text 'For You' is NOT visible.",
     "Recommendations visible; For You absent",
     "INTENTIONAL FAIL: Swap the two assertions (look for For You, deny Recommendations). Confirm both fail. Revert.",
     "None"],

    # ── Phase 3A — Auth Guards ────────────────────────────────────────────────
    ["TC-017","3A","Auth","/watchlist redirects logged-out user to login",
     "Playwright E2E","tests/e2e/auth.spec.js",
     "None",
     "Use a fresh browser context with NO saved auth state (logged-out context)",
     "1. Create new browser context with no storageState.\n2. Navigate to /watchlist.\n3. Wait for networkidle.\n4. Assert URL includes '/login' OR page contains 'sign in' text.",
     "Redirected to /login OR sign-in prompt visible",
     "INTENTIONAL FAIL: Assert URL includes '/dashboard'. Confirm fail. Revert.",
     "None"],

    ["TC-018","3A","Auth","/history redirects logged-out user to login",
     "Playwright E2E","tests/e2e/auth.spec.js",
     "None",
     "Logged-out browser context",
     "1. New logged-out context.\n2. Navigate to /history.\n3. Assert URL includes '/login' OR sign-in text visible.",
     "Redirected to /login OR sign-in prompt visible",
     "Same as TC-017 verification approach",
     "None"],

    ["TC-019","3A","Auth","/recommendations redirects logged-out user to login",
     "Playwright E2E","tests/e2e/auth.spec.js",
     "None",
     "Logged-out context",
     "1. New logged-out context.\n2. Navigate to /recommendations.\n3. Assert URL includes '/login' OR sign-in text visible.",
     "Redirected to /login OR sign-in prompt visible",
     "Same as TC-017",
     "None"],

    ["TC-020","3A","Auth","Logged-out /support shows sign-in link, not checkout button",
     "Playwright E2E","tests/e2e/auth.spec.js",
     "None",
     "Logged-out context",
     "1. New logged-out context.\n2. Navigate to /support.\n3. Assert text 'sign in to support' (case-insensitive) is visible.\n4. Assert no POST to /create-checkout-session in network intercept.",
     "'Sign in to support' visible; no checkout API call triggered",
     "WATCH: Run --headed. Confirm no checkout network call appears.\nINTENTIONAL FAIL: Look for 'Support CineMatch' button instead. Should fail. Revert.",
     "None"],

    ["TC-021","3A","Auth","Logged-out nav shows Sign In / Register links",
     "Playwright E2E","tests/e2e/auth.spec.js",
     "None",
     "Logged-out context",
     "1. New logged-out context.\n2. Navigate to /.\n3. Assert nav contains 'Sign In' or 'Sign in' text.",
     "Sign In link visible; no user avatar",
     "INTENTIONAL FAIL: Assert nav contains 'Sign Out'. Should fail. Revert.",
     "None"],

    # ── Phase 3B/C — Email Auth ────────────────────────────────────────────────
    ["TC-022","3B","Auth","Register form shows validation errors on empty submit",
     "Playwright E2E","tests/e2e/auth.spec.js",
     "None",
     "Logged-out context",
     "1. Navigate to /register.\n2. Leave all fields empty.\n3. Click Register/Sign Up button.\n4. Assert at least one error message element is visible.",
     "Validation error messages appear; form not submitted",
     "INTENTIONAL FAIL: Assert NO error messages exist. Should fail on empty submit. Revert.",
     "None"],

    ["TC-026","3C","Auth","Wrong password shows error message without crash",
     "Playwright E2E","tests/e2e/auth.spec.js",
     "Test email: any registered address. Wrong password: 'WrongPass9999!'",
     "Logged-out context; a registered email account exists",
     "1. Navigate to /login.\n2. Fill email with registered address.\n3. Fill password with 'WrongPass9999!'.\n4. Click Sign In.\n5. Assert error message visible.\n6. Assert URL is still /login.",
     "Error message shown; page stays on /login; no redirect",
     "INTENTIONAL FAIL: Assert error message is NOT visible. Should fail. Revert.",
     "None"],

    ["TC-027","3C","Auth","Correct credentials redirect to homepage with session active",
     "Playwright E2E","tests/e2e/auth.spec.js",
     "TEST_EMAIL and TEST_PASSWORD stored in .env file in qa-automation/",
     "Create a test account manually first (TC-023). Store email/password as TEST_EMAIL and TEST_PASSWORD in qa-automation/.env",
     "1. Navigate to /login.\n2. Fill email from process.env.TEST_EMAIL.\n3. Fill password from process.env.TEST_PASSWORD.\n4. Click Sign In.\n5. Wait for networkidle.\n6. Assert URL is '/' or '/browse'.\n7. Save auth state to tests/.auth/user.json for reuse.",
     "Redirected to homepage; logged-in; auth state saved to file",
     "INTENTIONAL FAIL: Use wrong password. Confirm test fails on URL assertion. Revert to correct password.",
     "Auth state saved — used by all subsequent tests that need login"],

    ["TC-028","3C","Auth","Session persists after browser page refresh",
     "Playwright E2E","tests/e2e/auth.spec.js",
     "Saved auth state from TC-027 (tests/.auth/user.json)",
     "Run TC-027 first to generate auth state",
     "1. Load context with storageState: 'tests/.auth/user.json'.\n2. Navigate to /.\n3. Reload page (page.reload()).\n4. Assert nav still shows user avatar or account element.",
     "User avatar/name still visible after reload; session persists",
     "INTENTIONAL FAIL: Load context without storageState. Assert avatar visible. Should fail. Revert.",
     "None"],

    # ── Phase 4 — Browse ──────────────────────────────────────────────────────
    ["TC-032","4","Browse","Browse page loads genre list successfully",
     "Playwright E2E","tests/e2e/browse.spec.js",
     "None",
     "None",
     "1. Navigate to /browse.\n2. Wait for networkidle.\n3. Assert genre button/chip elements count > 0.\n4. Assert at least one genre label matches known genres (Action, Drama, Comedy).",
     "Genre list renders with multiple options",
     "NETWORK CHECK: Add request listener for /genres API call. Confirm call is made on page load.\nINTENTIONAL FAIL: Assert genre count > 100. Should fail. Revert.",
     "None"],

    ["TC-033","4","Browse","Selecting a genre populates movie grid",
     "Playwright E2E","tests/e2e/browse.spec.js",
     "None",
     "None",
     "1. Navigate to /browse.\n2. Wait for genres to load.\n3. Click first available genre button.\n4. Wait 2000ms for movies to load.\n5. Count movie card elements.\n6. Assert count >= 4.",
     "Movie grid populates with 4+ cards after genre click",
     "NETWORK CHECK: Intercept /movies?genre=... call. Confirm it fires after genre click.\nINTENTIONAL FAIL: Assert count >= 200. Should fail. Revert.",
     "None"],

    ["TC-034","4","Browse","Pagination loads a different set of movies",
     "Playwright E2E","tests/e2e/browse.spec.js",
     "A genre with more than 20 movies (Action or Drama recommended)",
     "Genre already selected from TC-033",
     "1. Select a genre with many movies.\n2. Record title of first visible movie card.\n3. Click Next page button.\n4. Wait for load.\n5. Assert first movie title on page 2 differs from page 1 title.",
     "Different movies shown on page 2",
     "INTENTIONAL FAIL: Assert page 2 first title equals page 1 first title. Should fail for any multi-page genre. Revert.",
     "None"],

    ["TC-035","4","Browse","Movie card click navigates to /movies/[slug]",
     "Playwright E2E","tests/e2e/browse.spec.js",
     "None",
     "Genre selected, movies showing",
     "1. Click first movie card in browse grid.\n2. Wait for navigation.\n3. Assert page.url() contains '/movies/'.",
     "URL contains /movies/ after click",
     "INTENTIONAL FAIL: Assert URL contains '/people/'. Should fail. Revert.",
     "None"],

    # ── Phase 5 — Movie Detail ─────────────────────────────────────────────────
    ["TC-036","5","Movie Detail","Detail page shows title, year, genres, overview, cast, poster",
     "Playwright E2E","tests/e2e/movie-detail.spec.js",
     "Test movie: 'Inception' (known to exist in dataset)",
     "None",
     "1. Navigate to /movies/Inception.\n2. Assert movie title visible.\n3. Assert year text matches /\\d{4}/.\n4. Assert at least one genre tag visible.\n5. Assert overview/synopsis text length > 20 chars.\n6. Assert at least one cast member visible.\n7. Assert poster img element is visible and not broken.",
     "All 6 elements present and populated",
     "INTENTIONAL FAIL: Assert title text equals 'Avatar'. Should fail for Inception page. Revert.",
     "None"],

    ["TC-037","5","Movie Detail","Add to Watchlist button visible when logged in",
     "Playwright E2E","tests/e2e/movie-detail.spec.js",
     "Saved auth state from TC-027",
     "Run TC-027 first. Load storageState: 'tests/.auth/user.json'",
     "1. Load auth context.\n2. Navigate to /movies/Inception.\n3. Assert 'Add to Watchlist' button (or similar label) is visible.",
     "'Add to Watchlist' button present and visible",
     "LOGGED-OUT CHECK: Run same test without auth state. Confirm button is NOT visible to logged-out users.",
     "None"],

    ["TC-038","5","Movie Detail","Clicking Add to Watchlist changes button state and logs no errors",
     "Playwright E2E","tests/e2e/movie-detail.spec.js",
     "Auth state. Test movie: 'Inception'. SETUP: Ensure Inception is NOT already in watchlist before test.",
     "Load auth state. If Inception is in watchlist, remove it via UI first (or use API call to clean).",
     "1. Load auth context.\n2. Collect console errors.\n3. Navigate to /movies/Inception.\n4. Note initial button label.\n5. Click Add to Watchlist.\n6. Assert button label or state changes.\n7. Assert no new console errors after click.",
     "Button changes from 'Add' to 'Added' or 'Remove'; zero console errors",
     "INTENTIONAL FAIL: Assert button label does NOT change. Should fail after click. Revert.\nNETWORK CHECK: Intercept Supabase or API call on click. Confirm request is made.",
     "CLEANUP: After test, click Remove from Watchlist OR remove Inception from watchlist via Supabase API call so test starts clean next run."],

    ["TC-039","5","Movie Detail","Clicking a cast member navigates to /people/[slug]",
     "Playwright E2E","tests/e2e/movie-detail.spec.js",
     "Test movie: 'Inception'",
     "None",
     "1. Navigate to /movies/Inception.\n2. Locate first clickable cast member.\n3. Click them.\n4. Wait for navigation.\n5. Assert URL contains '/people/'.",
     "URL contains /people/ after cast click",
     "INTENTIONAL FAIL: Assert URL contains '/movies/' (not /people/). Should fail. Revert.",
     "None"],

    # ── Phase 6 — People ──────────────────────────────────────────────────────
    ["TC-040","6","People","Person page shows name, movies, and photo (or fallback)",
     "Playwright E2E","tests/e2e/people.spec.js",
     "Navigate from Inception cast (TC-039) or go directly to /people/Leonardo%20DiCaprio",
     "None",
     "1. Navigate to a people page (e.g. /people/Leonardo%20DiCaprio).\n2. Assert person name text is visible.\n3. Assert at least one movie title listed.\n4. Assert img element present (photo or fallback placeholder — no broken icon).",
     "Name, movies list, and image (or clean fallback) all visible",
     "INTENTIONAL FAIL: Assert person name equals 'Fake Person'. Should fail. Revert.",
     "None"],

    ["TC-041","6","People","Movie title links on people page navigate to /movies/[slug]",
     "Playwright E2E","tests/e2e/people.spec.js",
     "On a people page with movie list visible",
     "Navigate to people page first",
     "1. On people page, click first movie title link.\n2. Wait for navigation.\n3. Assert URL contains '/movies/'.",
     "URL contains /movies/ after click",
     "INTENTIONAL FAIL: Assert URL contains '/people/'. Should fail. Revert.",
     "None"],

    # ── Phase 7 — Search ──────────────────────────────────────────────────────
    ["TC-042","7","Search","Searching 'Inception' returns Inception in results",
     "Playwright E2E","tests/e2e/search.spec.js",
     "Search term: 'Inception'",
     "None",
     "1. Navigate to search page or find search input.\n2. Type 'Inception'.\n3. Wait 1500ms for results.\n4. Assert result list contains text 'Inception'.",
     "Inception appears in search results",
     "NETWORK CHECK: Intercept GET /search?query=Inception. Confirm call fires after typing.\nINTENTIONAL FAIL: Assert results contain 'ZZZNOTAMOVIE'. Should fail. Revert.",
     "None"],

    ["TC-043","7","Search","Fuzzy query 'Incepshun' returns non-empty results",
     "Playwright E2E","tests/e2e/search.spec.js",
     "Fuzzy query: 'Incepshun'",
     "None",
     "1. Type 'Incepshun' into search.\n2. Wait 1500ms.\n3. Assert result list count > 0.",
     "At least one result returned (fuzzy matching active)",
     "INTENTIONAL FAIL: Assert result count equals 0. Should fail for fuzzy search. Revert.",
     "None"],

    ["TC-044","7","Search","Query over 200 chars returns 400 error",
     "Playwright API","tests/e2e/search.spec.js",
     "201-char string: 'a'.repeat(201)",
     "None",
     "1. Build string of 201 'a' characters.\n2. Send GET /search?query={string} via APIRequestContext.\n3. Assert HTTP status === 400.",
     "HTTP 400; error code in response",
     "INTENTIONAL FAIL: Assert status === 200. Should fail. Revert.\nALSO: Send 200-char string, assert status 200. Proves boundary is exactly 200, not 201.",
     "None"],

    ["TC-045","7","Search","Clearing search input removes all results",
     "Playwright E2E","tests/e2e/search.spec.js",
     "Search results currently showing",
     "Type a search query first to get results",
     "1. Search for 'Inception' (results showing).\n2. Clear input (triple-click + Delete).\n3. Wait 500ms.\n4. Assert results container is empty or shows placeholder text.",
     "Results cleared; no ghost cards remain",
     "INTENTIONAL FAIL: Assert results still contain 'Inception' after clearing. Should fail. Revert.",
     "None"],

    # ── Phase 8 — Recommendations ─────────────────────────────────────────────
    ["TC-046","8","Recommendations","Empty input shows prompt to search for a film",
     "Playwright E2E","tests/e2e/recommendations.spec.js",
     "None",
     "None",
     "1. Navigate to /recommendations.\n2. Assert an empty-state message or prompt text is visible (no results grid).",
     "Empty state prompt shown; no recommendation cards",
     "INTENTIONAL FAIL: Assert recommendation cards count > 0 without input. Should fail. Revert.",
     "None"],

    ["TC-047","8","Recommendations","Recommendations load after searching for Avatar",
     "Playwright E2E","tests/e2e/recommendations.spec.js",
     "Movie title: 'Avatar'",
     "None",
     "1. Navigate to /recommendations.\n2. Type 'Avatar' in search input.\n3. Wait for autocomplete, click Avatar suggestion.\n4. Wait for recommendations to load (networkidle).\n5. Assert recommendation card count >= 5.",
     "5+ recommendation cards visible",
     "NETWORK CHECK: Intercept GET /recommend?movie=Avatar. Confirm call is made.\nINTENTIONAL FAIL: Assert count >= 50. Should fail. Revert.",
     "None"],

    ["TC-048","8","Recommendations","Recommended titles differ from the searched movie",
     "Playwright E2E","tests/e2e/recommendations.spec.js",
     "Avatar recommendations loaded (TC-047)",
     "TC-047 must pass first",
     "1. After Avatar recommendations load, collect all card title texts.\n2. Assert none of the collected titles equals 'Avatar'.",
     "No recommendation card has the same title as Avatar",
     "INTENTIONAL FAIL: Assert one of the titles IS 'Avatar'. Should fail (Avatar won't recommend itself). Revert.",
     "None"],

    ["TC-049","8","Recommendations","Clicking recommendation card navigates to /movies/[slug]",
     "Playwright E2E","tests/e2e/recommendations.spec.js",
     "Avatar recommendations loaded",
     "TC-047 must pass first",
     "1. Click first recommendation card.\n2. Wait for navigation.\n3. Assert URL contains '/movies/'.",
     "URL contains /movies/ after click",
     "INTENTIONAL FAIL: Assert URL contains '/people/'. Should fail. Revert.",
     "None"],

    # ── Phase 9 — Watchlist ───────────────────────────────────────────────────
    ["TC-050","9","Watchlist","Watchlist displays previously added movies",
     "Playwright E2E","tests/e2e/watchlist.spec.js",
     "Auth state (TC-027). SETUP: 'Inception' must be in watchlist before test runs.",
     "SETUP STEP: Run setup script that navigates to /movies/Inception and clicks Add to Watchlist. Alternatively, insert directly via Supabase API in beforeAll hook.",
     "1. Load auth context.\n2. Navigate to /watchlist.\n3. Assert 'Inception' text is visible in the list.",
     "'Inception' appears in watchlist",
     "STATE CHECK: Remove Inception from watchlist manually, run test, confirm it FAILS. Re-add it, confirm PASS. This proves the test is reading live data.",
     "CLEANUP: After test suite completes, remove Inception from watchlist via UI or Supabase API."],

    ["TC-051","9","Watchlist","Removing a movie from watchlist removes the card",
     "Playwright E2E","tests/e2e/watchlist.spec.js",
     "Auth state. SETUP: 'The Martian' must be in watchlist before test.",
     "SETUP STEP: Add 'The Martian' to watchlist in beforeEach hook (navigate to its movie page, click Add to Watchlist).",
     "1. Load auth context.\n2. Navigate to /watchlist.\n3. Find The Martian card.\n4. Click its Remove button.\n5. Wait 1000ms.\n6. Assert The Martian card is no longer visible.",
     "Card removed from DOM; no page crash",
     "INTENTIONAL FAIL: Assert The Martian IS still visible after removing. Should fail. Revert.\nSTATE CHECK: After test, navigate back to /watchlist. The Martian should not be there.",
     "CLEANUP: The Martian is removed by the test itself. No further cleanup needed."],

    ["TC-052","9","Watchlist","Watchlist data persists after browser page reload",
     "Playwright E2E","tests/e2e/watchlist.spec.js",
     "Auth state. SETUP: 'Interstellar' must be in watchlist.",
     "SETUP STEP: Add Interstellar to watchlist in beforeEach.",
     "1. Load auth context.\n2. Navigate to /watchlist.\n3. Assert Interstellar visible.\n4. Call page.reload().\n5. Wait for networkidle.\n6. Assert Interstellar still visible after reload.",
     "Interstellar persists after reload; data is in Supabase not just component state",
     "KILL TEST: Remove Interstellar from watchlist between step 3 and 5 using Supabase API. Reload. Test should FAIL (data gone proves it reads DB). Then restore.",
     "CLEANUP: Remove Interstellar from watchlist after test via Supabase API call."],

    # ── Phase 10 — History ────────────────────────────────────────────────────
    ["TC-053","10","History","/history page renders without crash in empty state",
     "Playwright E2E","tests/e2e/history.spec.js",
     "Auth state",
     "Load auth state",
     "1. Load auth context.\n2. Navigate to /history.\n3. Wait for networkidle.\n4. Assert page body text length > 10 (not a blank screen).\n5. Assert no JavaScript crash (console error listener active).",
     "Page renders; shows either movie list or empty-state message; no crash",
     "INTENTIONAL FAIL: Assert body text length === 0. Should fail for any rendered page. Revert.",
     "None"],

    ["TC-054","10","History","Viewed movies appear in browsing history",
     "Playwright E2E","tests/e2e/history.spec.js",
     "Auth state. SETUP: Navigate to /movies/Inception before checking history.",
     "Load auth state, then visit /movies/Inception to trigger history tracking",
     "1. Load auth context.\n2. Navigate to /movies/Inception (triggers history).\n3. Navigate to /history.\n4. Assert 'Inception' text visible in history list.",
     "Inception appears in history after being viewed",
     "INTENTIONAL FAIL: Assert history contains 'ZZZFAKEMOVIE'. Should fail. Revert.",
     "None"],

    # ── Phase 11 — Chat ───────────────────────────────────────────────────────
    ["TC-055","11","Chat","Chat responds to movie recommendation query",
     "Playwright E2E","tests/e2e/chat.spec.js",
     "enable_chat flag must be ON. GROQ_API_KEY must be set on backend.",
     "Confirm feature flag is enabled: GET /admin/feature-flags → enable_chat === true",
     "1. Navigate to chat page or find chat input.\n2. Type 'Recommend movies like The Dark Knight'.\n3. Click Send or press Enter.\n4. Wait up to 12000ms for response.\n5. Assert response text element is visible and not empty.",
     "AI response appears; response text length > 20 chars",
     "KILL TEST: POST to /admin/feature-flags to disable enable_chat. Run test. Confirm it catches the 503 error or disabled UI state. Restore flag.",
     "None"],

    ["TC-056","11","Chat","Chat response arrives within 10 seconds",
     "Playwright E2E","tests/e2e/chat.spec.js",
     "Same as TC-055",
     "Same as TC-055",
     "1. Record Date.now() before sending message.\n2. Send message.\n3. Wait for response element to become visible.\n4. Record Date.now() after response.\n5. Assert difference < 10000ms.",
     "Response time under 10 000ms",
     "INTENTIONAL FAIL: Change threshold to 1ms. Should always fail. Revert to 10000.",
     "None"],

    ["TC-058","11","Chat","Message over 1000 characters is rejected before sending",
     "Playwright E2E","tests/e2e/chat.spec.js",
     "String of 1001 'a' chars",
     "None",
     "1. Fill chat input with 'a'.repeat(1001).\n2. Click Send.\n3. Assert error message visible OR send button remains disabled.\n4. Assert no API call to /chat in network intercept.",
     "Message rejected; error shown or button disabled; no backend call",
     "INTENTIONAL FAIL: Fill with only 10 chars, assert error visible. Should fail (10 chars is valid). Revert.",
     "None"],

    ["TC-059","11","Chat","Whitespace-only message is rejected",
     "Playwright E2E","tests/e2e/chat.spec.js",
     "Spaces only: '     '",
     "None",
     "1. Fill chat input with 5 spaces.\n2. Click Send or press Enter.\n3. Collect all chat message elements.\n4. Assert none has empty/whitespace-only inner text.\n5. Assert no call to /chat in network intercept.",
     "No blank message in chat; no backend call made",
     "NETWORK CHECK: Add request listener for /chat. Confirm it does NOT fire on whitespace submit.",
     "None"],

    # ── Phase 12A — Support Auth Guard ────────────────────────────────────────
    ["TC-060","12A","Support","Logged-out /support shows sign-in link not checkout button",
     "Playwright E2E","tests/e2e/support.spec.js",
     "None",
     "Logged-out context (no auth state)",
     "1. New logged-out context.\n2. Navigate to /support.\n3. Assert 'sign in to support' text visible (case-insensitive).\n4. Assert no button with text 'Support CineMatch' or 'Pay' visible.",
     "Sign-in link visible; no checkout button",
     "LOGGED-IN CHECK: Run same test with auth state loaded. The checkout button SHOULD be visible. If it's not, the guard is broken in both directions.",
     "None"],

    ["TC-061","12A","Support","Sign-in link on support page navigates to /login",
     "Playwright E2E","tests/e2e/support.spec.js",
     "None",
     "Logged-out context",
     "1. New logged-out context.\n2. Navigate to /support.\n3. Click 'sign in to support' link.\n4. Wait for navigation.\n5. Assert URL includes '/login'.",
     "URL includes /login after clicking sign-in link",
     "INTENTIONAL FAIL: Assert URL includes '/register'. Should fail. Revert.",
     "None"],

    ["TC-062","12A","Support","No POST to /create-checkout-session for logged-out users",
     "Playwright E2E","tests/e2e/support.spec.js",
     "None",
     "Logged-out context; network request interception enabled",
     "1. New logged-out context.\n2. Register request listener: flag any request URL containing 'create-checkout-session'.\n3. Navigate to /support.\n4. Try clicking all visible buttons.\n5. Assert checkout flag is still false.",
     "Zero calls to /create-checkout-session from logged-out session",
     "WATCH: Run --headed. Open DevTools inside Playwright browser. Manually confirm no checkout call in Network tab.",
     "None"],

    # ── Phase 13 — Report Form ────────────────────────────────────────────────
    ["TC-079","13","Report","Empty report form shows validation errors on all required fields",
     "Playwright E2E","tests/e2e/report.spec.js",
     "None",
     "None",
     "1. Navigate to /report.\n2. Click Submit without filling any field.\n3. Assert at least one error message element is visible.",
     "Validation errors visible; form not submitted",
     "INTENTIONAL FAIL: Assert NO error messages are visible on empty submit. Should fail. Revert.",
     "None"],

    ["TC-080","13","Report","Subject shorter than 3 chars triggers length validation error",
     "Playwright E2E","tests/e2e/report.spec.js",
     "Subject input value: 'AB' (2 characters)",
     "None",
     "1. Navigate to /report.\n2. Select a valid category.\n3. Fill subject with 'AB'.\n4. Fill description with 50+ chars.\n5. Click Submit.\n6. Assert error message about subject length visible.",
     "Error message references 3-120 character requirement",
     "BOUNDARY CHECK: Fill subject with exactly 3 chars ('ABC'), submit. Test should PASS (no error). This proves 3 chars is valid.",
     "None"],

    ["TC-081","13","Report","Valid bug report without attachment submits successfully",
     "Playwright E2E","tests/e2e/report.spec.js",
     "Category: Bug. Subject: 'Playwright TC-081 auto-test'. Description: 'This is an automated QA test report submitted by Playwright to verify the full reporting pipeline is working correctly.'",
     "None",
     "1. Navigate to /report.\n2. Select Category = Bug.\n3. Fill Subject with 'Playwright TC-081 auto-test'.\n4. Fill Description with the 100+ char test string.\n5. Click Submit.\n6. Assert success message visible.",
     "Success message shown; no errors",
     "INTENTIONAL FAIL: Change expected success text to 'WRONG MESSAGE'. Should fail. Revert.\nKILL TEST: Take the backend offline or break the endpoint URL. Run. Confirm test catches the failure.",
     "DATA CREATED: A report row is added to Supabase reports table. Check admin panel to confirm, then optionally delete via Supabase dashboard."],

    ["TC-082","13","Report","Success message is clearly visible after valid submission",
     "Playwright E2E","tests/e2e/report.spec.js",
     "Same as TC-081 (runs in same test block)",
     "TC-081 setup",
     "1. After TC-081 submission succeeds.\n2. Assert visible text includes 'Thank you' or 'received'.",
     "Clear success message shown to user",
     "INTENTIONAL FAIL: Assert success message says 'Error'. Should fail. Revert.",
     "Same as TC-081"],

    ["TC-084","13A","Report","Report with image attachment submits successfully",
     "Playwright E2E","tests/e2e/report.spec.js",
     "FIXTURE FILE: qa-automation/tests/fixtures/test-image.jpg (create a small valid JPG, e.g. 1KB red pixel image)",
     "Place test-image.jpg in qa-automation/tests/fixtures/ before running.",
     "1. Navigate to /report.\n2. Select Category = Bug.\n3. Fill Subject: 'Playwright TC-084 attachment test'.\n4. Fill Description: 60+ chars.\n5. Use page.setInputFiles() to attach tests/fixtures/test-image.jpg.\n6. Click Submit.\n7. Assert success message visible.",
     "Success message shown; no file type error",
     "WRONG FILE CHECK: Swap the fixture to test.zip. Run. Confirm test FAILS on the error message assertion. Swap back.",
     "DATA CREATED: Report row in Supabase with has_attachment=true. Check admin panel to confirm."],

    ["TC-086","13A","Report","Unsupported file type (.zip) triggers error, not submission",
     "Playwright E2E","tests/e2e/report.spec.js",
     "FIXTURE FILE: qa-automation/tests/fixtures/test.zip (create an empty zip file)",
     "Place test.zip in qa-automation/tests/fixtures/ before running.",
     "1. Navigate to /report.\n2. Select Category = Bug.\n3. Fill Subject and Description with valid values.\n4. Attach tests/fixtures/test.zip.\n5. Click Submit.\n6. Assert error message about unsupported file type visible.\n7. Assert success message NOT visible.",
     "Error message shown referencing allowed file types; no submission",
     "INTENTIONAL FAIL: Assert success message IS visible after zip upload. Should fail. Revert.",
     "None — no data created (form rejected)"],

    # ── Phase 14A — Admin ─────────────────────────────────────────────────────
    ["TC-087","14A","Admin","Wrong admin credentials show error, no dashboard access",
     "Playwright E2E","tests/e2e/admin.spec.js",
     "Wrong credentials: username='hacker', password='wrongpass'",
     "None",
     "1. Navigate to /admin/login.\n2. Fill username: 'hacker'.\n3. Fill password: 'wrongpass'.\n4. Click Login.\n5. Assert error message visible.\n6. Assert URL does NOT include '/admin/dashboard' or '/admin' main page.",
     "Error shown; admin dashboard not accessible",
     "INTENTIONAL FAIL: Assert error message is NOT visible. Should fail on wrong credentials. Revert.",
     "None"],

    ["TC-088","14A","Admin","Admin dashboard loads after successful login",
     "Playwright E2E","tests/e2e/admin.spec.js",
     "Admin credentials stored in ADMIN_EMAIL env var. Auth state saved to tests/.auth/admin.json",
     "Run admin auth setup: npx playwright test tests/admin.setup.js (saves admin session to tests/.auth/admin.json)",
     "1. Load admin auth state from tests/.auth/admin.json.\n2. Navigate to /admin.\n3. Assert dashboard content is visible (not login page).",
     "Admin dashboard loads; not redirected to login",
     "INTENTIONAL FAIL: Load regular user auth state instead of admin. If admin gate works, should show access denied or redirect. Confirm test fails.",
     "None"],

    ["TC-089","14A","Admin","Admin dashboard renders Stats, Users, Reports, Feature Flags sections",
     "Playwright E2E","tests/e2e/admin.spec.js",
     "Admin auth state",
     "Load admin auth state",
     "1. Load admin context.\n2. Navigate to /admin.\n3. Assert 'Stats' section visible.\n4. Assert 'Users' or user table visible.\n5. Assert 'Reports' section visible.\n6. Assert feature flags section visible.",
     "All 4 sections rendered; no blank panels",
     "INTENTIONAL FAIL: Assert 'Payments History' section visible (doesn't exist). Should fail. Revert.",
     "None"],

    # ── Phase 14B — Admin Dashboard ───────────────────────────────────────────
    ["TC-090","14B","Admin","Stats section shows total_movies and total_genres > 0",
     "Playwright E2E","tests/e2e/admin.spec.js",
     "Admin auth state",
     "Load admin auth state",
     "1. Load admin context.\n2. Navigate to admin stats section.\n3. Extract total_movies number from page.\n4. Assert it is > 0.\n5. Extract total_genres and assert > 0.",
     "Both counts greater than zero",
     "INTENTIONAL FAIL: Assert total_movies > 999999. Should fail. Revert.",
     "None"],

    ["TC-091","14B","Admin","Users section displays at least 1 user profile row",
     "Playwright E2E","tests/e2e/admin.spec.js",
     "Admin auth state; at least 1 registered user exists",
     "Load admin auth state",
     "1. Load admin context.\n2. Navigate to users section.\n3. Count table row elements.\n4. Assert count >= 1.",
     "At least 1 user row in the table",
     "INTENTIONAL FAIL: Assert count >= 100. Should fail unless 100+ users. Revert.",
     "None"],

    ["TC-092","14B","Admin","Reports section shows TC-081 auto-test report",
     "Playwright E2E","tests/e2e/admin.spec.js",
     "Admin auth state. DEPENDENCY: TC-081 must have run first (creates the report).",
     "Run TC-081 before this test. The report 'Playwright TC-081 auto-test' must exist in Supabase.",
     "1. Load admin context.\n2. Navigate to reports section.\n3. Assert text 'Playwright TC-081 auto-test' is visible in the list.",
     "'Playwright TC-081 auto-test' report visible in admin panel",
     "STATE CHECK: Delete the TC-081 report from Supabase. Run this test. Should FAIL (report gone). Re-run TC-081 to recreate, then re-run TC-092. Confirms live DB read.",
     "None (the report data is left in place for visibility)"],

    ["TC-093","14B","Admin","Refresh button reloads reports list without full page reload",
     "Playwright E2E","tests/e2e/admin.spec.js",
     "Admin auth state; on reports section",
     "Load admin auth state",
     "1. Load admin context.\n2. Navigate to admin reports section.\n3. Record current page URL.\n4. Click 'Refresh' button.\n5. Wait 1500ms.\n6. Assert URL is unchanged (no full navigation).\n7. Assert reports still visible.",
     "URL unchanged after Refresh click; reports still showing",
     "INTENTIONAL FAIL: Assert URL changed after click. Should fail for in-place refresh. Revert.",
     "None"],

    # ── Phase 14C — Feature Flags ─────────────────────────────────────────────
    ["TC-094","14C","Feature Flags","Disabling enable_chat blocks /chat with FEATURE_DISABLED error",
     "Playwright API","tests/e2e/feature-flags.spec.js",
     "Backend awake; enable_chat currently ON",
     "Confirm enable_chat is ON via GET /admin/feature-flags before test",
     "1. POST /admin/feature-flags {flag:'enable_chat', enabled:false}.\n2. POST /chat with valid body.\n3. Assert response.error.code === 'FEATURE_DISABLED'.\n4. Assert HTTP status === 503.",
     "Chat returns 503 with FEATURE_DISABLED code when flag is off",
     "THIS TEST IS SELF-VALIDATING: It disables the flag and checks the result. If it passes, the flag system works. Run it — pass means correct.",
     "ALWAYS RESTORE: POST /admin/feature-flags {flag:'enable_chat', enabled:true} in afterEach hook."],

    ["TC-095","14C","Feature Flags","Re-enabling enable_chat restores /chat endpoint",
     "Playwright API","tests/e2e/feature-flags.spec.js",
     "enable_chat currently OFF (set by TC-094)",
     "TC-094 ran first and disabled the flag",
     "1. POST /admin/feature-flags {flag:'enable_chat', enabled:true}.\n2. GET /admin/feature-flags.\n3. Assert flags.enable_chat === true.",
     "enable_chat is true after re-enabling",
     "INTENTIONAL FAIL: Assert flags.enable_chat === false after enabling it. Should fail. Revert.",
     "None — flag restored to ON"],

    ["TC-096","14C","Feature Flags","Disabling enable_recommendations blocks /recommend endpoint",
     "Playwright API","tests/e2e/feature-flags.spec.js",
     "Backend awake; enable_recommendations currently ON",
     "Confirm flag is ON first",
     "1. POST /admin/feature-flags {flag:'enable_recommendations', enabled:false}.\n2. GET /recommend?movie=Avatar.\n3. Assert response.error.code === 'FEATURE_DISABLED'.\n4. Assert HTTP status === 503.",
     "Recommend endpoint returns 503 FEATURE_DISABLED when flag is off",
     "Same self-validating approach as TC-094.",
     "ALWAYS RESTORE: POST {flag:'enable_recommendations', enabled:true} in afterEach."],

    ["TC-097","14C","Feature Flags","Re-enabling enable_recommendations restores /recommend",
     "Playwright API","tests/e2e/feature-flags.spec.js",
     "enable_recommendations OFF (set by TC-096)",
     "TC-096 ran first",
     "1. POST /admin/feature-flags {flag:'enable_recommendations', enabled:true}.\n2. GET /admin/feature-flags.\n3. Assert flags.enable_recommendations === true.",
     "Flag is true after re-enabling",
     "INTENTIONAL FAIL: Assert false after enabling. Should fail. Revert.",
     "None — flag restored"],

    # ── Phase 16 — Edge Cases ─────────────────────────────────────────────────
    ["TC-104","16","Edge Cases","Non-existent movie URL shows 404 message, not blank screen",
     "Playwright E2E","tests/e2e/edge-cases.spec.js",
     "URL: /movies/this-movie-absolutely-does-not-exist-xyz-999",
     "None",
     "1. Navigate to /movies/this-movie-absolutely-does-not-exist-xyz-999.\n2. Wait for networkidle.\n3. Assert page body inner text length > 10.\n4. Assert page contains text matching /not found|404|does not exist/i.",
     "Meaningful 404 message shown; not a blank white screen",
     "BLANK CHECK: Assert body text length === 0. Should fail for any rendered page. Revert.\nINTENTIONAL FAIL: Assert page contains 'Welcome to CineMatch'. Should fail on 404 page. Revert.",
     "None"],

    ["TC-108","16","Edge Cases","Whitespace-only chat message is rejected before sending",
     "Playwright E2E","tests/e2e/edge-cases.spec.js",
     "Spaces string: '     ' (5 spaces)",
     "None",
     "1. Navigate to the chat page.\n2. Register request listener for /chat endpoint.\n3. Fill chat input with 5 spaces.\n4. Press Enter or click Send.\n5. Wait 1000ms.\n6. Assert /chat was NOT called (listener never fired).\n7. Assert no new message bubble with empty text appeared.",
     "No API call; no blank message; input rejected client-side",
     "NETWORK CHECK: This test relies on network interception to prove no server call was made. If the test passes but you see a /chat call in the listener, the assertion is wrong.",
     "None"],

    # ── Phase 17 — Cross-Browser (Playwright handles automatically) ───────────
    ["TC-109","17","Cross-Browser","Chrome: homepage + login + movie detail smoke test",
     "Playwright E2E","tests/e2e/homepage.spec.js",
     "Auth state; Playwright chromium project",
     "Defined in playwright.config.js projects array as {name:'Chrome', use:{browserName:'chromium'}}",
     "Playwright automatically runs all homepage and auth spec tests in Chromium. No additional code needed — just ensure the 'Chrome' project is in playwright.config.js.",
     "All homepage and movie detail tests pass in Chromium browser",
     "WATCH: Run npx playwright test --project=Chrome --headed. Watch Chromium browser perform tests.",
     "None"],

    ["TC-110","17","Cross-Browser","Firefox: homepage + login + movie detail smoke test",
     "Playwright E2E","tests/e2e/homepage.spec.js",
     "Auth state; Playwright firefox project",
     "Defined in playwright.config.js as {name:'Firefox', use:{browserName:'firefox'}}",
     "Playwright automatically runs the same spec files in Firefox. Run npx playwright test --project=Firefox.",
     "All homepage and movie detail tests pass in Firefox browser",
     "WATCH: Run npx playwright test --project=Firefox --headed to watch Firefox run the tests.",
     "None"],

    ["TC-111","17","Cross-Browser","Mobile (WebKit/Safari): homepage + login + movie detail",
     "Playwright E2E","tests/e2e/homepage.spec.js",
     "Auth state; Playwright webkit project with iPhone viewport",
     "Defined in playwright.config.js as {name:'Mobile Safari', use:{...devices['iPhone 14']}}",
     "Playwright runs all tests at iPhone 14 viewport (390px) using WebKit engine. Run npx playwright test --project='Mobile Safari'.",
     "All tests pass at mobile viewport; touch interactions work; no layout overflow",
     "WATCH: Run npx playwright test --project='Mobile Safari' --headed. Watch mobile viewport browser perform tests.",
     "None"],
]

# ══════════════════════════════════════════════════════════════════════════════
# SETUP GUIDE CONTENT
# ══════════════════════════════════════════════════════════════════════════════
SETUP_SECTIONS = [
    ("WHY PLAYWRIGHT?", [
        "Playwright is the definitive choice for CineMatch automation. It is the ONLY tool that covers all of:",
        "  - Backend API testing (no browser needed, via APIRequestContext)",
        "  - Browser E2E testing (Chromium, Firefox, WebKit in one tool)",
        "  - Auth state persistence (login once, reuse session across 72 tests)",
        "  - Network request interception (verify API calls, not just UI state)",
        "  - Mobile viewport simulation (iPhone 14, 375px — for TC-106, 107, 111)",
        "  - Auto-screenshot + video capture on test failure",
        "  - HTML report with timeline: npx playwright show-report",
        "  - Codegen: record your browser clicks → auto-generate test code",
        "",
        "Cypress: does NOT support Safari/WebKit, no multi-tab, weaker network intercept.",
        "Selenium: complex setup, slowest, no modern features. Skip entirely.",
        "Mocha+Chai: not a browser tool. Playwright's APIRequestContext replaces it.",
    ]),
    ("PREREQUISITES", [
        "1. Node.js 18+ installed (check: node --version)",
        "2. npm 9+ installed (check: npm --version)",
        "3. A test account created manually (TC-023): email + password stored in .env",
        "4. Backend awake at your-backend.onrender.com (first request may take 40s)",
        "5. Fixture files created:",
        "   qa-automation/tests/fixtures/test-image.jpg  (any valid JPG, even 1KB)",
        "   qa-automation/tests/fixtures/test.zip        (any zip file, even empty)",
    ]),
    ("INSTALLATION (run once)", [
        "cd c:\\Users\\PC\\Desktop\\VSC\\Projects\\Cinematch",
        "mkdir qa-automation",
        "cd qa-automation",
        "npm init -y",
        "npm install --save-dev @playwright/test dotenv",
        "npx playwright install chromium firefox webkit",
        "",
        "Create qa-automation/.env file:",
        "  TEST_EMAIL=your-test-account@email.com",
        "  TEST_PASSWORD=YourTestPassword123",
        "  BACKEND_URL=https://your-backend.onrender.com",
        "  FRONTEND_URL=https://your-app.vercel.app",
        "",
        "Add qa-automation/.env to .gitignore — never commit credentials.",
    ]),
    ("AUTH STATE SETUP (run once, re-run when session expires)", [
        "The auth state setup logs in once and saves the browser session to a file.",
        "All tests that need login load from this file — no repeated logins.",
        "",
        "Create qa-automation/tests/auth.setup.js:",
        "  const { chromium } = require('@playwright/test');",
        "  require('dotenv').config();",
        "  (async () => {",
        "    const browser = await chromium.launch();",
        "    const page = await browser.newPage();",
        "    await page.goto(process.env.FRONTEND_URL + '/login');",
        "    await page.fill('[type=email]', process.env.TEST_EMAIL);",
        "    await page.fill('[type=password]', process.env.TEST_PASSWORD);",
        "    await page.click('[type=submit]');",
        "    await page.waitForURL('**/*');",
        "    await page.context().storageState({ path: 'tests/.auth/user.json' });",
        "    await browser.close();",
        "    console.log('Auth state saved to tests/.auth/user.json');",
        "  })();",
        "",
        "Run it:  node tests/auth.setup.js",
        "Add tests/.auth/ to .gitignore.",
    ]),
    ("PLAYWRIGHT CONFIG (playwright.config.js)", [
        "module.exports = {",
        "  testDir: './tests/e2e',",
        "  timeout: 30000,",
        "  retries: 1,",
        "  reporter: [['html'], ['list']],",
        "  use: {",
        "    baseURL: process.env.FRONTEND_URL,",
        "    screenshot: 'only-on-failure',",
        "    video: 'retain-on-failure',",
        "  },",
        "  projects: [",
        "    { name: 'Chrome',        use: { browserName: 'chromium' } },",
        "    { name: 'Firefox',       use: { browserName: 'firefox'  } },",
        "    { name: 'Mobile Safari', use: { ...devices['iPhone 14'] } },",
        "  ],",
        "};",
    ]),
    ("RUNNING TESTS", [
        "# Run all tests (API + E2E, all browsers):",
        "  npx playwright test",
        "",
        "# Run API tests only (fast, no browser):",
        "  npx playwright test tests/e2e/health.spec.js tests/e2e/search.spec.js tests/e2e/feature-flags.spec.js",
        "",
        "# Run single spec file:",
        "  npx playwright test tests/e2e/homepage.spec.js",
        "",
        "# Run single test case by TC ID:",
        "  npx playwright test --grep TC-008",
        "",
        "# Watch browser run (best for learning):",
        "  npx playwright test --headed",
        "",
        "# Slow motion — see every step clearly:",
        "  npx playwright test --headed --slowMo=800",
        "",
        "# Run Chrome only:",
        "  npx playwright test --project=Chrome",
        "",
        "# Run mobile only:",
        "  npx playwright test --project='Mobile Safari'",
        "",
        "# Open HTML report after run:",
        "  npx playwright show-report",
        "",
        "# Record your browser clicks to generate test code:",
        "  npx playwright codegen https://your-app.vercel.app",
    ]),
    ("HOW TO VALIDATE AUTOMATION IS ACCURATE (Golden Rules)", [
        "RULE 1 — Intentional Failure Test (most important):",
        "  Every test must be able to FAIL. Before trusting a test:",
        "  a) Run it — should PASS.",
        "  b) Change one assertion to a wrong value (e.g. 200 → 999).",
        "  c) Run again — must FAIL with a clear error message.",
        "  d) Revert. If step (c) did not fail, the test proves nothing.",
        "",
        "RULE 2 — Network Interception:",
        "  Don't just check the UI — check that the API call was made.",
        "  Add: page.on('request', req => { if (req.url().includes('/recommend')) called = true })",
        "  Assert called === true after the action.",
        "",
        "RULE 3 — Feature Flag Kill Switch:",
        "  The best E2E validation: disable a feature via API, run the test,",
        "  confirm the test catches the broken state. Restore. Tests that",
        "  pass when features are broken are false positives.",
        "",
        "RULE 4 — Data State Verification:",
        "  For watchlist/report tests: add data, assert it appears, remove it,",
        "  assert it disappears. Both directions must work.",
        "",
        "RULE 5 — npx playwright show-report:",
        "  After any failure, open the HTML report. It shows the exact step",
        "  that failed, a screenshot at that moment, and the full error stack.",
        "  Compare screenshot to what you saw in manual testing.",
    ]),
    ("ADDING NEW TEST CASES TO CI", [
        "Question: After all tests pass, can new test cases be added to CI?",
        "Answer: YES — and it requires ZERO changes to the CI YAML.",
        "",
        "How it works:",
        "  1. GitHub Actions CI runs: npx playwright test",
        "  2. Playwright scans tests/e2e/ for all *.spec.js files automatically.",
        "  3. To add a new test: add a test() block to any existing .spec.js file,",
        "     OR create a new file: tests/e2e/my-new-feature.spec.js",
        "  4. Push to main. CI picks it up on the next run. Done.",
        "",
        "To run only critical tests in CI (for speed):",
        "  Tag tests with @smoke: test('TC-008: @smoke health check', ...)",
        "  CI command: npx playwright test --grep @smoke",
        "  Full regression run weekly: npx playwright test (all tags)",
        "",
        "GitHub Actions YAML snippet (add to .github/workflows/ci.yml):",
        "  playwright-tests:",
        "    runs-on: ubuntu-latest",
        "    steps:",
        "      - uses: actions/checkout@v3",
        "      - uses: actions/setup-node@v3",
        "        with: { node-version: '18' }",
        "      - run: cd qa-automation && npm install",
        "      - run: cd qa-automation && npx playwright install --with-deps chromium",
        "      - run: cd qa-automation && npx playwright test --project=Chrome",
        "        env:",
        "          TEST_EMAIL: ${{ secrets.TEST_EMAIL }}",
        "          TEST_PASSWORD: ${{ secrets.TEST_PASSWORD }}",
        "          FRONTEND_URL: https://your-app.vercel.app",
        "          BACKEND_URL: https://your-backend.onrender.com",
        "      - uses: actions/upload-artifact@v3",
        "        if: failure()",
        "        with:",
        "          name: playwright-report",
        "          path: qa-automation/playwright-report/",
    ]),
]

# ══════════════════════════════════════════════════════════════════════════════
# WORKBOOK BUILDER
# ══════════════════════════════════════════════════════════════════════════════

wb = Workbook()
wb.remove(wb.active)

def apply_header_row(ws, row, text, bg, fg, size=13, merge_to=None):
    ws.row_dimensions[row].height = 36 if size >= 13 else 26
    if merge_to:
        ws.merge_cells(f"A{row}:{merge_to}{row}")
    c = ws.cell(row=row, column=1, value=text)
    c.fill = F(bg); c.font = Font(name="Calibri", size=size, bold=True, color=fg)
    c.alignment = Alignment(horizontal="center", vertical="center")

def write_data_row(ws, row_num, values, col_configs, row_bg, border_color="C5D9EE"):
    """Write a data row. col_configs: list of (width_ignored, halign, bg_override_or_None)"""
    ws.row_dimensions[row_num].height = 14  # minimum; Excel auto-expands with wrap
    for ci, (val, cfg) in enumerate(zip(values, col_configs), start=1):
        h, bg_ov = cfg
        cell = ws.cell(row=row_num, column=ci, value=val)
        cell.fill = F(bg_ov if bg_ov else row_bg)
        cell.font = Fn(size=9)
        cell.alignment = Alignment(horizontal=h, vertical="top", wrap_text=True)
        cell.border = Bd(border_color)

# ────────────────────────────────────────────────────────────────────────────
# SHEET 1: DASHBOARD
# ────────────────────────────────────────────────────────────────────────────

ws_dash = wb.create_sheet("Dashboard")
ws_dash.sheet_view.showGridLines = False
ws_dash.sheet_properties.tabColor = "1F4E79"

for col, w in zip("ABCDEFGHIJKLM", [3,14,42,10,10,10,10,10,12,3,3,3,3]):
    ws_dash.column_dimensions[col].width = w

r = 1
ws_dash.row_dimensions[r].height = 8; r += 1

ws_dash.row_dimensions[r].height = 44
ws_dash.merge_cells(f"B{r}:I{r}")
c = ws_dash.cell(row=r, column=2, value="CINEMATCH — QA MASTER TEST PLAN")
c.fill = F(HDR_BG); c.font = Font(name="Calibri", size=22, bold=True, color="FFFFFF")
c.alignment = Alignment(horizontal="center", vertical="center"); r += 1

ws_dash.row_dimensions[r].height = 22
ws_dash.merge_cells(f"B{r}:I{r}")
c = ws_dash.cell(row=r, column=2, value="your-app.vercel.app  |  Backend: your-backend.onrender.com  |  Automation: Playwright")
c.fill = F("2E5FA3"); c.font = Font(name="Calibri", size=10, color="FFFFFF")
c.alignment = Alignment(horizontal="center", vertical="center"); r += 1

ws_dash.row_dimensions[r].height = 18
ws_dash.merge_cells(f"B{r}:I{r}")
c = ws_dash.cell(row=r, column=2, value="v5.0  |  Manual: Sheet 2  |  Automation: Sheet 3  |  Setup Guide: Sheet 4")
c.fill = F(PH_BG); c.font = Font(name="Calibri", size=9, color=PH_FG)
c.alignment = Alignment(horizontal="center", vertical="center"); r += 2

# Summary counts
from collections import defaultdict
manual_counts  = defaultdict(int)
auto_counts    = defaultdict(int)
for tc in MANUAL: manual_counts[tc[1]] += 1
for tc in AUTO:   auto_counts[tc[1]]   += 1

all_phases = list(PHASES.keys())
dash_hdrs = ["Phase", "Phase Name", "Manual TCs", "Auto TCs", "Total",
             "Pass","Fail","Blocked","% Done"]
hdr_r = r
ws_dash.row_dimensions[hdr_r].height = 22
for ci, h in enumerate(dash_hdrs, start=2):
    c = ws_dash.cell(row=hdr_r, column=ci, value=h)
    c.fill = F(HDR_BG); c.font = Fn(bold=True, color="FFFFFF", size=9)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = Bd("1F4E79")
r += 1

for ri, pk in enumerate(all_phases):
    ws_dash.row_dimensions[r].height = 17
    bg = EVEN_BG if ri % 2 == 0 else ODD_BG
    m  = manual_counts.get(pk, 0)
    a  = auto_counts.get(pk, 0)
    row_vals = [pk, PHASES[pk].split("—")[-1].strip(), m, a, m+a, "","","",""]
    for ci, v in enumerate(row_vals, start=2):
        c = ws_dash.cell(row=r, column=ci, value=v)
        c.fill = F(bg); c.font = Fn(size=9)
        c.alignment = Alignment(horizontal="center" if ci not in (3,) else "left",
                                 vertical="center")
        c.border = Bd("C5D9EE")
    # % formula based on Pass col (col 7 = index 7 from B = col H)
    col_pass  = get_column_letter(7)
    col_total = get_column_letter(7-2)  # Total col = col E = 6th col from A = col 6
    # Total is col F (ci=6 from B means column=6+1=7... let me use direct col letters)
    # B=2,C=3,D=4,E=5,F=6,G=7,H=8,I=9,J=10
    # Phase=B(2), Name=C(3), Manual=D(4), Auto=E(5), Total=F(6), Pass=G(7), Fail=H(8), Blocked=I(9), %=J(10)
    c_pct = ws_dash.cell(row=r, column=10)
    c_pct.value = f'=IF(F{r}=0,"",G{r}/F{r})'
    c_pct.number_format = "0%"
    c_pct.fill = F(bg); c_pct.font = Fn(size=9, bold=True)
    c_pct.alignment = Alignment(horizontal="center", vertical="center")
    c_pct.border = Bd("C5D9EE")
    r += 1

# Totals
ws_dash.row_dimensions[r].height = 20
ws_dash.merge_cells(f"B{r}:C{r}")
c = ws_dash.cell(row=r, column=2, value="TOTAL")
c.fill = F(HDR_BG); c.font = Fn(bold=True, color="FFFFFF", size=9)
c.alignment = Alignment(horizontal="center", vertical="center")
ws_dash.cell(row=r, column=3).fill = F(HDR_BG)
dr = hdr_r + 1
for ci in range(4, 11):
    cl = get_column_letter(ci)
    c2 = ws_dash.cell(row=r, column=ci)
    if ci == 10:
        c2.value = f'=IF(F{r}=0,"",G{r}/F{r})'
        c2.number_format = "0%"
    else:
        c2.value = f"=SUM({cl}{dr}:{cl}{r-1})"
    c2.fill = F(HDR_BG); c2.font = Fn(bold=True, color="FFFFFF", size=9)
    c2.alignment = Alignment(horizontal="center", vertical="center")
    c2.border = Bd("1F4E79")
r += 2

# Legend
ws_dash.merge_cells(f"B{r}:I{r}")
c = ws_dash.cell(row=r, column=2, value="SHEET GUIDE")
c.fill = F(PH_BG); c.font = Font(name="Calibri", size=10, bold=True, color=PH_FG)
c.alignment = Alignment(horizontal="left", vertical="center"); r += 1

legend = [
    ("Sheet 2 — Manual Test Cases",     "~39 TCs requiring human judgment: OAuth, Stripe payment, Telegram verification, visual checks, Supabase dashboard"),
    ("Sheet 3 — Automation Test Cases", "~72 TCs automated by Playwright. Each row includes what the script does, how to verify it works, and data cleanup."),
    ("Sheet 4 — Automation Setup Guide","Step-by-step: install Playwright, save auth state, run tests, add tests to CI, validate accuracy."),
    ("Status values (both sheets)",     "Pass | Fail | Blocked (dependency failed) | Skip (deliberate, note reason) | Not Tested"),
    ("Priority (manual only)",          "Critical = must pass before release  |  High = should pass  |  Medium = nice to have  |  Low = setup only"),
    ("Golden Rule (automation)",        "Change one assertion to a wrong value. Run. It MUST fail with a clear error. Revert. If it didn't fail, the test proves nothing."),
]
for i, (k, v) in enumerate(legend):
    ws_dash.row_dimensions[r].height = 20
    ws_dash.merge_cells(f"B{r}:C{r}")
    c = ws_dash.cell(row=r, column=2, value=k)
    bg = EVEN_BG if i%2==0 else ODD_BG
    c.fill = F(bg); c.font = Fn(bold=True, size=9)
    c.alignment = Alignment(horizontal="left", vertical="center")
    c.border = Bd()
    ws_dash.cell(row=r, column=3).fill = F(bg)
    ws_dash.merge_cells(f"D{r}:I{r}")
    c2 = ws_dash.cell(row=r, column=4, value=v)
    c2.fill = F(bg); c2.font = Fn(size=9)
    c2.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    c2.border = Bd()
    for ci in range(5,10):
        ws_dash.cell(row=r, column=ci).fill = F(bg)
    r += 1

# ────────────────────────────────────────────────────────────────────────────
# SHEET 2: MANUAL TEST CASES
# ────────────────────────────────────────────────────────────────────────────

ws_m = wb.create_sheet("Manual Test Cases")
ws_m.sheet_view.showGridLines = False
ws_m.sheet_properties.tabColor = "2E5FA3"
ws_m.freeze_panes = "A3"

M_COLS = [
    ("TC ID",            9,  "center"),
    ("Phase",            8,  "center"),
    ("Category",        15,  "left"),
    ("Test Case Title", 38,  "left"),
    ("Priority",        11,  "center"),
    ("Preconditions",   30,  "left"),
    ("Test Steps",      55,  "left"),
    ("Expected Result", 46,  "left"),
    ("Actual Result",   32,  "left"),
    ("Status",          13,  "center"),
    ("Tester",          14,  "left"),
    ("Date Tested",     13,  "center"),
    ("Notes / Defect",  28,  "left"),
]
for ci, (_, w, _) in enumerate(M_COLS, 1):
    ws_m.column_dimensions[get_column_letter(ci)].width = w

ws_m.row_dimensions[1].height = 32
ws_m.merge_cells(f"A1:{get_column_letter(len(M_COLS))}1")
c = ws_m.cell(row=1, column=1, value="CINEMATCH — MANUAL TEST CASES  |  Phases 0, 3B/D, 12B-E, 13 (Telegram), 14D, 15, 16 (Visual)")
c.fill = F(HDR_BG); c.font = Font(name="Calibri", size=12, bold=True, color=HDR_FG)
c.alignment = Alignment(horizontal="center", vertical="center")

ws_m.row_dimensions[2].height = 28
for ci, (label, _, halign) in enumerate(M_COLS, 1):
    c = ws_m.cell(row=2, column=ci, value=label)
    c.fill = F("2E5FA3"); c.font = Fn(bold=True, color="FFFFFF", size=9)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = Bd("1F4E79")

m_dv = DataValidation(type="list", formula1='"Pass,Fail,Blocked,Skip,Not Tested"', allow_blank=True)
ws_m.add_data_validation(m_dv)

m_row = 3; prev_ph = None; m_counter = 0
for tc in MANUAL:
    tc_id, ph, cat, title, prio, pre, steps, exp = tc
    if ph != prev_ph:
        prev_ph = ph; m_counter = 0
        ws_m.row_dimensions[m_row].height = 18
        ws_m.merge_cells(f"A{m_row}:{get_column_letter(len(M_COLS))}{m_row}")
        c = ws_m.cell(row=m_row, column=1, value=PHASES[ph])
        c.fill = F(PH_BG); c.font = Font(name="Calibri", size=9, bold=True, italic=True, color=PH_FG)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        c.border = Bd_med(); m_row += 1

    bg = ODD_BG if m_counter%2==0 else EVEN_BG; m_counter += 1
    nl = steps.count("\n")
    ws_m.row_dimensions[m_row].height = max(55, nl*13+18)

    vals = [tc_id, f"Ph {ph}", cat, title, prio, pre, steps, exp, "", "", "", "", ""]
    configs = [
        ("center", None), ("center","F0F0F0"), ("left",None), ("left",None),
        ("center", PRIO_BG.get(prio, bg)),
        ("left",None), ("left",None), ("left",None),
        ("left",EMPTY_BG), ("center",EMPTY_BG), ("left",EMPTY_BG),
        ("center",EMPTY_BG), ("left",EMPTY_BG),
    ]
    for ci, (val, (h, bg_ov)) in enumerate(zip(vals, configs), 1):
        cell = ws_m.cell(row=m_row, column=ci, value=val)
        cell.fill = F(bg_ov if bg_ov else bg)
        bold = ci in (1, 4)
        fg_c = PRIO_FG.get(prio, "1A1A2E") if ci==5 else ("1F4E79" if ci==1 else "1A1A2E")
        cell.font = Fn(bold=bold, color=fg_c, size=9)
        cell.alignment = Alignment(horizontal=h, vertical="top", wrap_text=True)
        cell.border = Bd()

    m_dv.add(ws_m.cell(row=m_row, column=10))
    m_row += 1

# ────────────────────────────────────────────────────────────────────────────
# SHEET 3: AUTOMATION TEST CASES
# ────────────────────────────────────────────────────────────────────────────

ws_a = wb.create_sheet("Automation Test Cases")
ws_a.sheet_view.showGridLines = False
ws_a.sheet_properties.tabColor = "145A32"
ws_a.freeze_panes = "A4"

A_COLS = [
    ("TC ID",                  9,  "center"),
    ("Phase",                  8,  "center"),
    ("Category",              14,  "left"),
    ("Test Case Title",       35,  "left"),
    ("Tool",                  16,  "center"),
    ("Test File",             30,  "left"),
    ("Test Data / Fixtures",  28,  "left"),
    ("Setup Before Test",     28,  "left"),
    ("What the Script Does",  52,  "left"),
    ("Expected Result",       38,  "left"),
    ("How to Verify Accuracy",44,  "left"),
    ("Data Cleanup",          28,  "left"),
    ("Status",                13,  "center"),
    ("Last Run Date",         13,  "center"),
    ("Notes",                 24,  "left"),
]
for ci, (_, w, _) in enumerate(A_COLS, 1):
    ws_a.column_dimensions[get_column_letter(ci)].width = w

ws_a.row_dimensions[1].height = 32
ws_a.merge_cells(f"A1:{get_column_letter(len(A_COLS))}1")
c = ws_a.cell(row=1, column=1, value="CINEMATCH — AUTOMATION TEST CASES  |  Tool: Playwright  |  72 Test Cases  |  DO NOT run steps manually — these are executed by Playwright")
c.fill = F(AUTO_HDR); c.font = Font(name="Calibri", size=11, bold=True, color=AUTO_FG)
c.alignment = Alignment(horizontal="center", vertical="center")

ws_a.row_dimensions[2].height = 20
ws_a.merge_cells(f"A2:{get_column_letter(len(A_COLS))}2")
c = ws_a.cell(row=2, column=1,
    value="Run: cd qa-automation && npx playwright test   |   Watch: add --headed --slowMo=800   |   Report: npx playwright show-report   |   Single test: --grep TC-008")
c.fill = F("145A32"); c.font = Font(name="Calibri", size=9, color="ABEBC6")
c.alignment = Alignment(horizontal="center", vertical="center")

ws_a.row_dimensions[3].height = 28
for ci, (label, _, _) in enumerate(A_COLS, 1):
    c = ws_a.cell(row=3, column=ci, value=label)
    c.fill = F("145A32"); c.font = Fn(bold=True, color="FFFFFF", size=9)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = Bd("0B5345")

a_dv = DataValidation(type="list", formula1='"Pass,Fail,Blocked,Skip,Not Run"', allow_blank=True)
ws_a.add_data_validation(a_dv)

a_row = 4; prev_ph_a = None; a_counter = 0
for tc in AUTO:
    (tc_id, ph, cat, title, tool, tfile, tdata, setup,
     what_does, expected, verify, cleanup) = tc

    if ph != prev_ph_a:
        prev_ph_a = ph; a_counter = 0
        ws_a.row_dimensions[a_row].height = 18
        ws_a.merge_cells(f"A{a_row}:{get_column_letter(len(A_COLS))}{a_row}")
        c = ws_a.cell(row=a_row, column=1, value=PHASES[ph])
        c.fill = F(AUTO_PH); c.font = Font(name="Calibri", size=9, bold=True, italic=True, color=AUTO_PFG)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        c.border = Bd_med("A9DFBF"); a_row += 1

    bg = ODD_BG if a_counter%2==0 else "F0FAF4"; a_counter += 1
    nl = max(what_does.count("\n"), verify.count("\n"), setup.count("\n"))
    ws_a.row_dimensions[a_row].height = max(60, nl*13+22)

    # Tool badge colour
    tool_bg = "D6EAF8" if "API" in tool else PW_BG

    vals = [tc_id, f"Ph {ph}", cat, title, tool, tfile, tdata,
            setup, what_does, expected, verify, cleanup, "", "", ""]
    configs = [
        ("center",None),("center","F0F0F0"),("left",None),("left",None),
        ("center",tool_bg),("left","FDFEFE"),("left","FDFEFE"),
        ("left","FFFDE7"),   # setup = yellow tint
        ("left",None),      # what script does
        ("left","EBF5FB"),  # expected = blue tint
        ("left","FEF9E7"),  # verify = amber tint
        ("left","FDEDEC"),  # cleanup = red tint
        ("center",EMPTY_BG),("center",EMPTY_BG),("left",EMPTY_BG),
    ]
    for ci, (val, (h, bg_ov)) in enumerate(zip(vals, configs), 1):
        cell = ws_a.cell(row=a_row, column=ci, value=val)
        cell.fill = F(bg_ov if bg_ov else bg)
        bold = ci in (1, 4)
        fg_c = PW_FG if ci==5 else ("1F4E79" if ci==1 else "1A1A2E")
        cell.font = Fn(bold=bold, color=fg_c, size=9)
        cell.alignment = Alignment(horizontal=h, vertical="top", wrap_text=True)
        cell.border = Bd("A9DFBF" if bg_ov and "F0F" not in (bg_ov or "") else "C5D9EE")

    a_dv.add(ws_a.cell(row=a_row, column=13))
    a_row += 1

# ────────────────────────────────────────────────────────────────────────────
# SHEET 4: AUTOMATION SETUP GUIDE
# ────────────────────────────────────────────────────────────────────────────

ws_g = wb.create_sheet("Automation Setup Guide")
ws_g.sheet_view.showGridLines = False
ws_g.sheet_properties.tabColor = "7D3C98"
ws_g.column_dimensions["A"].width = 3
ws_g.column_dimensions["B"].width = 110

g_row = 1
ws_g.row_dimensions[g_row].height = 8; g_row += 1

ws_g.row_dimensions[g_row].height = 38
ws_g.merge_cells(f"A{g_row}:B{g_row}")
c = ws_g.cell(row=g_row, column=1, value="CINEMATCH — AUTOMATION SETUP GUIDE  |  Playwright  |  Read this before running any automated tests")
c.fill = F("4A235A"); c.font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
c.alignment = Alignment(horizontal="center", vertical="center"); g_row += 1

for sec_title, lines in SETUP_SECTIONS:
    ws_g.row_dimensions[g_row].height = 24
    ws_g.merge_cells(f"A{g_row}:B{g_row}")
    c = ws_g.cell(row=g_row, column=1, value=sec_title)
    c.fill = F("7D3C98"); c.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    g_row += 1

    for i, line in enumerate(lines):
        ws_g.row_dimensions[g_row].height = 15
        ws_g.merge_cells(f"A{g_row}:B{g_row}")
        is_code = line.startswith("  ") or "=>" in line or "npx" in line or ".js" in line
        c = ws_g.cell(row=g_row, column=1, value=line if line else " ")
        bg = "F5EEF8" if i%2==0 else ODD_BG
        c.fill = F("F0F0F0" if is_code else bg)
        c.font = Font(name="Courier New" if is_code else "Calibri",
                      size=8 if is_code else 9, color="1A1A2E")
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
        g_row += 1

    ws_g.row_dimensions[g_row].height = 6; g_row += 1

# ── Print setup ────────────────────────────────────────────────────────────
from openpyxl.worksheet.page import PageMargins
for ws in [ws_dash, ws_m, ws_a, ws_g]:
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5)
ws_m.page_setup.orientation = "landscape"; ws_m.page_setup.fitToWidth = 1; ws_m.print_title_rows = "1:2"
ws_a.page_setup.orientation = "landscape"; ws_a.page_setup.fitToWidth = 1; ws_a.print_title_rows = "1:3"

# ── Save ────────────────────────────────────────────────────────────────────
OUT = r"c:\Users\PC\Desktop\VSC\Projects\Cinematch\CineMatch_QA_Test_Plan.xlsx"
wb.save(OUT)
print(f"Saved: {OUT}")
print(f"  Manual TCs:    {len(MANUAL)}")
print(f"  Automation TCs:{len(AUTO)}")
print(f"  Total TCs:     {len(MANUAL)+len(AUTO)}")
