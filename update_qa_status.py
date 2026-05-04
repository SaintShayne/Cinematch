#!/usr/bin/env python3
"""
update_qa_status.py

Runs the full Playwright suite (JSON reporter), then stamps the
"Status" and "Last Run Date" columns in CineMatch_QA_Test_Plan.xlsx
for every automation TC row.

Usage:
    python update_qa_status.py

Requirements:  pip install openpyxl   (already installed if generate_qa_plan.py works)
"""

import json
import re
import subprocess
import sys
from collections import defaultdict
from datetime import date
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

# ── Paths ──────────────────────────────────────────────────────────────────────

ROOT   = Path(__file__).parent
XLSX   = ROOT / "CineMatch_QA_Test_Plan.xlsx"
QA_DIR = ROOT / "qa-automation"
TODAY  = date.today().strftime("%Y-%m-%d")

# ── Cell styles ────────────────────────────────────────────────────────────────

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

STYLE = {
    "Pass":    (fill("D5F5E3"), Font(name="Calibri", size=9, bold=True, color="1E8449")),
    "Fail":    (fill("FFD0D0"), Font(name="Calibri", size=9, bold=True, color="922B21")),
    "Blocked": (fill("FFE0B2"), Font(name="Calibri", size=9, bold=True, color="BF360C")),
}
DEFAULT_STYLE = (fill("F5F8FA"), Font(name="Calibri", size=9, color="1A1A2E"))

DATE_FONT = Font(name="Calibri", size=9, color="5D6D7E")
CENTER    = Alignment(horizontal="center", vertical="top", wrap_text=False)

# ── Playwright ─────────────────────────────────────────────────────────────────

def run_playwright():
    print("Running Playwright (JSON reporter)…")
    proc = subprocess.run(
        "npx playwright test --reporter=json",
        cwd=str(QA_DIR),
        capture_output=True,
        text=True,
        shell=True,
    )
    # Progress bar goes to stderr; clean JSON goes to stdout.
    if not proc.stdout.strip():
        print("ERROR: No JSON output from Playwright.")
        print("stderr:", proc.stderr[-2000:])
        sys.exit(1)
    return proc.stdout


# ── Parse results ──────────────────────────────────────────────────────────────

def _collect_specs(node, out):
    """Recursively pull every spec object out of the nested suite tree."""
    for spec in node.get("specs", []):
        out.append(spec)
    for child in node.get("suites", []):
        _collect_specs(child, out)


def parse_results(raw_json: str) -> dict[str, str]:
    """Return {TC_ID: 'Pass'|'Fail'} aggregated across all projects."""
    data = json.loads(raw_json)

    specs = []
    for suite in data.get("suites", []):
        _collect_specs(suite, specs)

    # A TC may appear in multiple projects (chromium, mobile, api).
    # Aggregate: Pass only if every run passed.
    by_tc: dict[str, list[bool]] = defaultdict(list)
    for spec in specs:
        title = spec.get("title", "")
        m = re.match(r"(TC-\d+)", title)
        if not m:
            continue
        by_tc[m.group(1)].append(bool(spec.get("ok", False)))

    return {tc: ("Pass" if all(oks) else "Fail") for tc, oks in by_tc.items()}


# ── Update xlsx ────────────────────────────────────────────────────────────────

def update_xlsx(results: dict[str, str]) -> int:
    if not XLSX.exists():
        print(f"ERROR: {XLSX} not found. Run python generate_qa_plan.py first.")
        sys.exit(1)

    wb = openpyxl.load_workbook(XLSX)

    if "Automation Test Cases" not in wb.sheetnames:
        print("ERROR: 'Automation Test Cases' sheet not found in the workbook.")
        sys.exit(1)

    ws = wb["Automation Test Cases"]
    updated = 0

    for row in ws.iter_rows(min_row=4):
        tc_cell = row[0]          # Column A — TC ID
        val = tc_cell.value
        if not val or not str(val).startswith("TC-"):
            continue              # Phase header row or empty — skip

        tc_id = str(val).strip()
        status = results.get(tc_id)

        # Column 13 = Status  (0-indexed: row[12])
        # Column 14 = Last Run Date  (0-indexed: row[13])
        status_cell = row[12]
        date_cell   = row[13]

        if status:
            st_fill, st_font = STYLE.get(status, DEFAULT_STYLE)
            status_cell.value     = status
            status_cell.fill      = st_fill
            status_cell.font      = st_font
            status_cell.alignment = CENTER
            date_cell.value       = TODAY
            date_cell.font        = DATE_FONT
            date_cell.alignment   = CENTER
            updated += 1
        elif status_cell.value in (None, "", "Not Run"):
            # TC exists in sheet but wasn't in this run — leave as-is
            pass

    # ── Dashboard summary banner ───────────────────────────────────────────────
    if "Dashboard" in wb.sheetnames:
        ws_dash = wb["Dashboard"]
        total  = len(results)
        passed = sum(1 for v in results.values() if v == "Pass")
        failed = total - passed
        # Find the cell that says "Last automation run" and write next to it.
        for row in ws_dash.iter_rows():
            for cell in row:
                if cell.value and "last automation run" in str(cell.value).lower():
                    # Write date in the cell immediately to the right
                    right = ws_dash.cell(row=cell.row, column=cell.column + 1)
                    right.value = TODAY
                    break

        # Try to find pass/fail count cells by looking for "Passed" label
        for row in ws_dash.iter_rows():
            for cell in row:
                v = str(cell.value or "").strip().lower()
                if v == "passed":
                    ws_dash.cell(row=cell.row, column=cell.column + 1).value = passed
                elif v == "failed":
                    ws_dash.cell(row=cell.row, column=cell.column + 1).value = failed
                elif v in ("total automation", "automation total"):
                    ws_dash.cell(row=cell.row, column=cell.column + 1).value = total

    wb.save(XLSX)
    return updated


# ── Main ───────────────────────────────────────────────────────────────────────

def main():
    verbose = "--verbose" in sys.argv or "-v" in sys.argv

    raw = run_playwright()
    results = parse_results(raw)

    passed = sum(1 for v in results.values() if v == "Pass")
    failed = len(results) - passed
    print(f"Parsed {len(results)} TC results — {passed} Pass / {failed} Fail")

    if verbose:
        print("\nTC ID → Status mapping:")
        for tc_id in sorted(results):
            status = results[tc_id]
            icon = "✓" if status == "Pass" else "✗"
            print(f"  {icon}  {tc_id}  →  {status}")
        print()

    updated = update_xlsx(results)
    print(f"Stamped {updated} rows in {XLSX.name} with status + date {TODAY}.")
    print("Open the file (or npx playwright show-report for the HTML report).")


if __name__ == "__main__":
    main()
